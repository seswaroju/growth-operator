"""CSV/XLSX extraction + column mapping (MVP-078) — the structured bulk-import stage.

Reads an uploaded price sheet, maps its columns to catalog fields **once per source signature**
(header-tuple hash, remembered in `tenant_settings` so the same sheet maps automatically next time),
and writes one `import_rows` row per data row (`raw` original + `normalized` mapped + `flags`). It
advances the batch `extracting → extracted` (or `failed`, resumable). No vision/LLM — that is the
photo path (MVP-077, gated). Per-org (RLS).
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.ingestion.state import BatchState
from core.ingestion.storage import ImportBlobStore, default_store
from core.tenancy.repository import set_org_context

# Source header (lowercased) → catalog field. Everything unmatched becomes an attribute.
_FIELD_ALIASES: dict[str, str] = {
    "name": "title", "title": "title", "item": "title", "product": "title",
    "sku": "sku", "code": "sku", "item code": "sku",
    "price": "base_price_minor", "mrp": "base_price_minor", "rate": "base_price_minor",
    "amount": "base_price_minor", "selling price": "base_price_minor",
    "description": "description", "desc": "description", "details": "description",
}
_MAPPED_FIELDS = {"title", "sku", "base_price_minor", "description"}
_MONEY = re.compile(r"[^\d.]")


class ExtractionFailed(Exception):
    """Extraction failed; the batch has been moved to `failed` (resumable) with the reason."""


def _signature(headers: list[str]) -> str:
    return hashlib.sha256("".join(h.strip().lower() for h in headers).encode()).hexdigest()[:16]


def auto_map(headers: list[str]) -> dict[str, str]:
    """Best-effort header → catalog-field map; unmatched headers map to ``attr:<header>``."""
    out: dict[str, str] = {}
    for h in headers:
        key = h.strip().lower()
        out[h] = _FIELD_ALIASES.get(key, f"attr:{h.strip()}")
    return out


def _to_minor(value: str) -> int | None:
    cleaned = _MONEY.sub("", value or "")
    if not cleaned:
        return None
    try:
        return int(round(float(cleaned) * 100))
    except ValueError:
        return None


def _parse(source_kind: str, data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    if source_kind == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(header_row)]
        data_rows = [
            {headers[i]: ("" if c is None else str(c)) for i, c in enumerate(r) if i < len(headers)}
            for r in rows_iter if any(c is not None for c in r)
        ]
        return headers, data_rows
    # CSV (default)
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
    headers = list(reader.fieldnames or [])
    return headers, [dict(r) for r in reader]


def apply_mapping(
    raw: dict[str, Any], mapping: dict[str, str]
) -> tuple[dict[str, Any], list[str], float]:
    """Map a raw row to a normalized catalog dict + flags + a coarse confidence."""
    normalized: dict[str, Any] = {"attributes": {}}
    flags: list[str] = []
    for src, val in raw.items():
        target = mapping.get(src, f"attr:{src}")
        text_val = "" if val is None else str(val).strip()
        if target == "base_price_minor":
            minor = _to_minor(text_val)
            if minor is None and text_val:
                flags.append(f"unparsed_price:{src}")
            normalized["base_price_minor"] = minor
        elif target in _MAPPED_FIELDS:
            normalized[target] = text_val
        else:  # attr:<name>
            normalized["attributes"][target.removeprefix("attr:")] = text_val
    if not normalized.get("title"):
        flags.append("missing_title")  # required — the row can't load without it
    confidence = 1.0 if not flags else round(max(0.3, 1.0 - 0.2 * len(flags)), 2)
    return normalized, flags, confidence


def _mapping_key(signature: str) -> str:
    return f"ingest.mapping.{signature}"


async def load_saved_mapping(
    session: AsyncSession, org_id: UUID, signature: str
) -> dict[str, str] | None:
    """The remembered column map for this source signature, if any (org-scoped tenant_settings).

    Stored directly (not via the settings service, which only knows registered config keys).
    """
    row = (await session.execute(
        text("SELECT value FROM tenant_settings WHERE org_id = :o AND key = :k "
             "ORDER BY version DESC LIMIT 1"),
        {"o": str(org_id), "k": _mapping_key(signature)})).scalar_one_or_none()
    if row is None:
        return None
    value = json.loads(row) if isinstance(row, str) else row
    return value if isinstance(value, dict) and value else None


async def save_mapping(
    session: AsyncSession, org_id: UUID, signature: str, mapping: dict[str, str]
) -> None:
    """Remember this source's column map so the same sheet maps automatically next time."""
    key = _mapping_key(signature)
    await session.execute(
        text("DELETE FROM tenant_settings WHERE org_id = :o AND key = :k"),
        {"o": str(org_id), "k": key})
    await session.execute(
        text("INSERT INTO tenant_settings (org_id, key, value) "
             "VALUES (:o, :k, CAST(:v AS jsonb))"),
        {"o": str(org_id), "k": key, "v": json.dumps(mapping)})


async def extract_batch(
    session: AsyncSession, org_id: UUID, batch_id: UUID,
    store: ImportBlobStore | None = None,
) -> int:
    """Extract one batch's uploaded sheet into `import_rows`. Returns the row count. Advances the
    batch state; on any failure moves it to `failed` (resumable) and re-raises."""
    from core.ingestion.service import transition

    store = store or default_store()
    await set_org_context(session, org_id)
    batch = (await session.execute(
        text("SELECT source_kind, storage_ref FROM import_batches WHERE id = :id AND org_id = :o"),
        {"id": str(batch_id), "o": str(org_id)})).mappings().first()
    if batch is None:
        raise KeyError(f"unknown import batch {batch_id}")
    await transition(session, org_id, batch_id, BatchState.extracting)
    try:
        data = await store.load(batch["storage_ref"])
        if data is None:
            raise ValueError("uploaded file not found in the blob store")
        headers, rows = _parse(str(batch["source_kind"]), data)
        signature = _signature(headers)
        mapping = await load_saved_mapping(session, org_id, signature) or auto_map(headers)
        # clear any prior extraction (a retry re-extracts cleanly)
        await session.execute(
            text("DELETE FROM import_rows WHERE batch_id = :b"), {"b": str(batch_id)})
        for seq, raw in enumerate(rows):
            normalized, flags, confidence = apply_mapping(raw, mapping)
            await session.execute(
                text("INSERT INTO import_rows "
                     "(org_id, batch_id, seq, raw, normalized, confidence, flags, state) "
                     "VALUES (:o, :b, :seq, CAST(:raw AS jsonb), CAST(:norm AS jsonb), "
                     "CAST(:conf AS jsonb), CAST(:flags AS jsonb), 'extracted')"),
                {"o": str(org_id), "b": str(batch_id), "seq": seq, "raw": json.dumps(raw),
                 "norm": json.dumps(normalized), "conf": json.dumps(confidence),
                 "flags": json.dumps(flags)})
        await save_mapping(session, org_id, signature, mapping)
        await session.execute(
            text("UPDATE import_batches SET row_count = :n WHERE id = :b"),
            {"n": len(rows), "b": str(batch_id)})
        await transition(session, org_id, batch_id, BatchState.extracted)
        return len(rows)
    except ExtractionFailed:
        raise
    except Exception as exc:  # parse/mapping error → mark failed (resumable), surface the reason
        await transition(session, org_id, batch_id, BatchState.failed)
        await session.execute(
            text("UPDATE import_batches SET error = :e WHERE id = :b"),
            {"e": f"extraction failed: {exc}"[:500], "b": str(batch_id)})
        raise ExtractionFailed(str(exc)) from exc
